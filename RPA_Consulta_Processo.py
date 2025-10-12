"""
Automação Jurídica - Consulta TJSP com IA e Logs

Instalação:
pip install -r requirements.txt
playwright install
"""

import logging
from datetime import datetime
from transformers import pipeline
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
import os
import re
import win32com.client as win32

# CONFIGURAÇÃO DO LOGGING
os.makedirs("logs", exist_ok=True)
data_atual = datetime.now().strftime("%Y%m%d_%H%M%S")
log_arquivo = f"logs/execucao_{data_atual}.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_arquivo, encoding="utf-8"),
        logging.StreamHandler()
    ]
)

logging.info("===== INICIANDO EXECUÇÃO DO ROBÔ JURÍDICO =====")

# CONFIGURAÇÃO DO PIPELINE DE IA
try:
    qa = pipeline("question-answering", model="pierreguillou/bert-base-cased-squad-v1.1-portuguese")
    logging.info("Modelo de IA carregado com sucesso (HuggingFace - BERT Português).")
except Exception as e:
    logging.exception("Erro ao carregar modelo de IA: %s", e)
    raise

# Enviar e-mail via Outlook
def enviar_email_outlook(destinatario, assunto, corpo, anexos=None):
    outlook = win32.Dispatch('outlook.application')
    email = outlook.CreateItem(0)
    email.To = destinatario
    email.Subject = assunto
    email.Body = corpo
    if anexos:
        for anexo in anexos:
            email.Attachments.Add(anexo)
    email.Send()
    print("E-mail enviado com sucesso pelo Outlook!")

    
# CARREGAMENTO DA PLANILHA
arquivo = "Processos.xlsx"
if not os.path.exists(arquivo):
    logging.error(f"Arquivo {arquivo} não encontrado.")
    raise FileNotFoundError(f"O arquivo {arquivo} não foi encontrado no diretório atual.")

wb = load_workbook(arquivo)
ws = wb.active
logging.info(f"Planilha '{arquivo}' carregada com sucesso. Aba ativa: {ws.title}")

# EXECUÇÃO PRINCIPAL DO ROBÔ
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()

    for row in ws.iter_rows(min_row=2):
        nome_parte = str(row[0].value).strip() if row[0].value else "Parte_Desconhecida"
        doc_contraparte = str(row[1].value).strip() if row[1].value else None
        if not doc_contraparte:
            continue

        logging.info(f"Consultando processo: {doc_contraparte}")

        try:

            # 🔹 Cria pasta da pessoa (baseada na planilha)
            nome_limpo = re.sub(r'[\\/*?:"<>|]', "_", nome_parte)
            pasta_destino = os.path.join("Processos", f"{nome_limpo}_{doc_contraparte}")
            os.makedirs(pasta_destino, exist_ok=True)

            # Acessa o site do TJSP
            page.goto("https://esaj.tjsp.jus.br/cpopg/open.do", timeout=15000)
            page.select_option("#cbPesquisa", "DOCPARTE")
            page.locator("#campo_DOCPARTE").fill(doc_contraparte)
            page.locator("#botaoConsultarProcessos").click()
            page.wait_for_timeout(5000)

            # Salva a página HTML
            nome_html = os.path.join(pasta_destino, f"{doc_contraparte}_tjsp.html")
            # with open(nome_html, "w", encoding="utf-8") as f:
            #     f.write(page.content())
            # logging.info(f"HTML salvo com sucesso em: {nome_html}")

            # Lê e converte o HTML para texto
            with open(nome_html, "r", encoding="utf-8") as f:
                soup = BeautifulSoup(f, "html.parser")
                contexto = soup.get_text()
            logging.info("Arquivo HTML carregado e processado para texto.")

            # Perguntas para a IA
            perguntas = [
                "Qual o número da ação?",
                "Qual o tipo da ação?",
                "Qual o valor da ação?",
                "Qual a comarca?",
                "Qual o nome do juiz?",
                "Qual o último andamento do processo?"
            ]

            # Loop para perguntas
            for pergunta in perguntas:
                try:
                    resposta = qa(question=pergunta, context=contexto)
                    resposta_texto = resposta["answer"]
                    logging.info(f"Pergunta: {pergunta} | Resposta: {resposta_texto}")

                    if "número da ação" in pergunta.lower():
                        row[2].value = resposta_texto   # Coluna C
                    elif "tipo da ação" in pergunta.lower():
                        row[3].value = resposta_texto   # Coluna D
                    elif "valor" in pergunta.lower():
                        row[4].value = resposta_texto   # Coluna E
                    elif "comarca" in pergunta.lower():
                        row[5].value = resposta_texto   # Coluna F
                    elif "juiz" in pergunta.lower():
                        row[6].value = resposta_texto   # Coluna G
                    elif "andamento" in pergunta.lower():
                        row[7].value = resposta_texto   # Coluna H

                except Exception as e:
                    logging.warning(f"Erro ao responder '{pergunta}' no processo {doc_contraparte}: {e}")

            wb.save("Processos.xlsx")
            logging.info(f"Planilha atualizada com sucesso para o processo {doc_contraparte}.")

            print(os.path.abspath(nome_html))
            enviar_email_outlook(
            destinatario="equipe@exemplo.com",
            assunto= f"Atualização do processo TJSP {nome_limpo}",
            corpo="Olá, equipe!\n\nSegue em anexo o resultado da automação.\n\nAtenciosamente,\nRobô Jurídico",
            anexos=[os.path.abspath(nome_html)]
            )

        except Exception as e:
            logging.exception(f"Erro ao processar o processo {doc_contraparte}: {e}")

    browser.close()

logging.info("===== EXECUÇÃO FINALIZADA COM SUCESSO =====")
logging.info(f"Arquivo de log salvo em: {log_arquivo}")